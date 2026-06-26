from pathlib import Path

from openpyxl import load_workbook

DROP_SHEETS = ("foto", "fotograf", "imagen")
KEEP_LABELS = {
    "marca", "modelo", "serie", "capacidad", "cliente", "rotulo",
    "fecha", "ot", "o.t", "tecnico", "area usuaria",
    "descripcion", "cumple", "no cumple", "no aplica",
    "observacion", "observaciones", "reparacion", "mantencion",
    "de baja", "estado", "analisis",
}
FREE_TEXT = (
    "inspeccion visual", "prueba de funcionamiento", "desarme",
    "procedimiento", "repuestos o accesorios requeridos",
)


def norm(value):
    text = str(value or "").strip().lower()
    return (text.replace("á", "a").replace("é", "e").replace("í", "i")
                .replace("ó", "o").replace("ú", "u"))


def clone_sheet_only(workbook, sheet):
    for name in list(workbook.sheetnames):
        if name != sheet.title:
            del workbook[name]
    sheet.title = "Informe Tecnico"


def best_sheet(workbook):
    candidates = []
    for sheet in workbook.worksheets:
        name = norm(sheet.title)
        if any(word in name for word in DROP_SHEETS):
            continue
        filled = sum(1 for row in sheet.iter_rows() for cell in row if cell.value)
        candidates.append((filled, sheet))
    return sorted(candidates, key=lambda item: item[0], reverse=True)[0][1]


def clear_right_of_labels(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            if norm(cell.value) in KEEP_LABELS:
                limit = min(sheet.max_column, cell.column + 3)
                for target in row[cell.column:limit]:
                    if norm(target.value) not in KEEP_LABELS:
                        target.value = None


def clear_marks_and_notes(sheet):
    for row in sheet.iter_rows():
        row_text = " ".join(norm(cell.value) for cell in row if cell.value)
        if "registro fotograf" in row_text:
            for cell in row:
                cell.value = None
            continue
        for cell in row:
            text = norm(cell.value)
            if text in ("x", "/"):
                cell.value = None
            if any(section in text for section in FREE_TEXT) and ":" in text:
                cell.value = None
            if "jefe de area" in text or "tecnico especializado" in text:
                cell.value = None


def clean_template(input_path, output_path):
    wb = load_workbook(input_path)
    sheet = best_sheet(wb)
    clone_sheet_only(wb, sheet)
    clear_right_of_labels(sheet)
    clear_marks_and_notes(sheet)
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
